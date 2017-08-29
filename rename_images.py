# coding:utf-8
import os
from openpyxl import load_workbook


class RenameImages(object):
    def __init__(self, image_path, excel_path):
        self.image_path = image_path
        self.excel_path = excel_path

    def get_replace_rules(self):
        wb = load_workbook(self.excel_path)
        ws = wb.active

        rules = {}
        for index in range(1, ws.max_row + 1):
            rules[str(ws.cell(row=index, column=1).value).strip()] = str(ws.cell(row=index, column=2).value).strip()
        return rules

    def process(self):
        rules = self.get_replace_rules()

        for image in os.listdir(self.image_path):
            if os.path.isfile(os.path.join(self.image_path, image)) and \
                    image.endswith(('.jpg', '.jpeg', '.png')):
                sn, index = image.split('_(')
                if sn in rules:
                    new_image = rules[sn] + '_(' + index
                    print 'Renamed {} => {}'.format(image, new_image)
                    os.rename(os.path.join(self.image_path, image), os.path.join(self.image_path, new_image))


def main():
    image_path = r'E:\task\replace-images-name'.decode('utf-8')
    excel_path = r'C:\Users\Administrator\Desktop\商品sn命名.xlsx'.decode('utf-8')

    reImage = RenameImages(image_path, excel_path)
    reImage.process()


if __name__ == "__main__":
    main()
