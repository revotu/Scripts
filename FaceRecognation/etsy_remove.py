# -*- coding:utf-8 -*-
import os
from openpyxl import load_workbook


def remove_images(images_dir, remove_file):
    with open(remove_file) as fp:
        for line in fp:
            brand, sn = line.strip().split('\t')
            for file in os.listdir(os.path.join(images_dir, brand)):
                if file.startswith(sn):
                    print file
                    os.remove(os.path.join(images_dir, brand, file))

def remove_excel_data(remove_file, excel_data):
    wb = load_workbook(excel_data)
    ws = wb.active
    with open(remove_file) as f:
        for line in f:
            brand, sn = line.strip().split('\t')
            for row_index in range(1, ws.max_row+1):
                if ws.cell(row=row_index, column=2).value and \
                        ws.cell(row=row_index, column=4).value and \
                        ws.cell(row=row_index, column=2).value.startswith(sn) and \
                        ws.cell(row=row_index, column=4).value == brand:

                    ws.cell(row=row_index, column=1).value = None
                    ws.cell(row=row_index, column=2).value = None
                    ws.cell(row=row_index, column=3).value = None
                    ws.cell(row=row_index, column=4).value = None
                    ws.cell(row=row_index, column=5).value = None
                    ws.cell(row=row_index, column=6).value = None
                    print row_index,brand,sn
                    break
    wb.save(excel_data)

def main():
    images_dir = r'D:\work\dev-crawler\avarsha\images\etsy'
    remove_file = r'D:\work\dev-crawler\remove.txt'
    excel_data = r'D:\work\dev-crawler\etsy-data.xlsx'
    #remove_images(images_dir, remove_file)
    remove_excel_data(remove_file, excel_data)


if __name__ == "__main__":
    main()