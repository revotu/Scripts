import os.path
import urllib

import openpyxl
from openpyxl.reader.excel import load_workbook

def youtubeChecker(videoID):
    checkUrl = 'http://www.youtube.com/oembed?url=http://www.youtube.com/watch?v=%s' % (videoID)
    try:
        response = urllib.urlopen(checkUrl)
        status = response.getcode()
        if status == 200:
            return True
        else:
            return False
    except:
        return False


def checkVideoLink(videoLink):
    videoID = videoLink.split('/')[-1]
    return youtubeChecker(videoID)

if __name__ == "__main__":
    wb = load_workbook('enki-video.xlsx')
    ws = wb.active
    for index in range(1,ws.get_highest_row() + 1):
        videoLink = ws.cell(row = index,column = 2).value
        ws.cell(row = index,column = 3).value = checkVideoLink(videoLink)
        print index,videoLink
        wb.save('enki-video.xlsx')
#     wb = load_workbook('enki-video.xlsx')
#     ws = wb.active
#     f = open('enki-video-error-link.txt',a)
#     for index in range(1,ws.get_highest_row() + 1):
#         status = ws.cell(row = index,column = 3).value
#         link = ws.cell(row = index,column = 1).value
#         if status == 'False':
#             f.write('http://www.enkivillage.com/%s.html' % (link))
#     f.close()