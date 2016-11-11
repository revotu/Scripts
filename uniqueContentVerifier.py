import urllib
import urllib2
import logging
import md5
import os
import json
import string
import time
import traceback
import utils
from lxml import html
from lxml.html import clean

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from pyvirtualdisplay import Display

from openpyxl import load_workbook

logging.basicConfig()
selenium_logger = logging.getLogger(
    'selenium.webdriver.remote.remote_connection')
selenium_logger.setLevel(logging.WARNING)
logger = logging.getLogger('UniqueContentVerifier')


def RemovePunctuation(str_txt):
    new_str = ""
    for ch in str_txt:
        if string.punctuation.find(ch) == -1:
            new_str += ch
    return new_str

def SnippetSeenQuery(raw_snippet, raw_query):
    raw_snippet = raw_snippet.lower()
    raw_snippet = raw_snippet.replace("\n", "")
    raw_snippet = RemovePunctuation(raw_snippet)
    raw_query = raw_query.lower()
    raw_query = RemovePunctuation(raw_query)
    query_words = raw_query.split(" ")
    for i in range(len(query_words) - 1):
        sub_query = "%s %s" % (query_words[i], query_words[i + 1])
        if raw_snippet.find(sub_query) == -1:
            return False
    return True

def IsNotSeenInSearchAPI(raw_query):
    query = urllib.urlencode({'q': raw_query})
    url = 'https://ajax.googleapis.com/ajax/services/search/web?v=1.0&%s' % query
    request = urllib2.Request(url)
    response = urllib2.urlopen(request)
    search_results = response.read()
    results = json.loads(search_results)
    data = results['responseData']
    if data is None:
        return "NOTSURE"
    hits = data['results']
    if len(hits) == 0:
        return "NOTSURE"

    for h in hits:
        raw_snippet = utils.StripHtmlTag(h['content'])
        if SnippetSeenQuery(raw_snippet, raw_query):
            return False
    return True

class UniqueContentVerifier(object):
    def __init__(self, lang='en', headless=False):
        # Linux use headless mode
        self.display = None
        if os.name == 'posix' and headless:
            self.display = Display(visible=0, size=(800, 600))
            self.display.start()

        self.WAIT_TIME = 120
        options = webdriver.ChromeOptions()
        options.add_argument('--lang=%s' % lang)
        self.driver = webdriver.Chrome(chrome_options=options)
        self.driver.implicitly_wait(self.WAIT_TIME)
        logger.info('Log in to google successfully.')

    def close(self):
        if self.display != None:
            self.display.stop()
        self.driver.quit()
        logger.info('Goodbye, Google.')

    def is_content_unique_in_general_search(self, raw_query):
        query = urllib.urlencode({'q': raw_query})
        url = 'https://www.google.com/?gws_rd=ssl#%s' % (query)

        try:
            self.driver.get(url)
            time.sleep(15)
            wait = WebDriverWait(self.driver, self.WAIT_TIME)
            snippets = wait.until(EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, 'span.st')))
        except:
            logger.error('Wait timeout, google again.')
#             self.driver.get(url)
#             wait = WebDriverWait(self.driver, self.WAIT_TIME)
#             snippets = wait.until(EC.presence_of_all_elements_located(
#                 (By.CSS_SELECTOR, 'span.st')))

        page_source = self.driver.page_source.encode('utf-8')

        try:
            for snippet in snippets:
                if SnippetSeenQuery(snippet.text, raw_query) is True:
                    return False
        except:
            logger.error('Exception in fetching google snippet.')
            with open('google.html', 'w') as fp:
                fp.write(page_source)
            return False

        return True

    def is_content_unique_in_site_search(self, raw_query):
        return self.is_content_unique_in_general_search("site:facebook.com " + raw_query)

    # Please check the result:
    # - NOTSURE
    # - True
    # - False
    def is_content_unique(self, raw_query):
        # unique_in_search_API = IsNotSeenInSearchAPI(raw_query)
        # if unique_in_search_API != "NOTSURE":
        #    return unique_in_search_API
        is_unique_in_general_search = self.is_content_unique_in_general_search(raw_query)
        if is_unique_in_general_search is False:
            return False
        if self.is_content_unique_in_site_search(raw_query) is False:
            return False
        return True


def main():
    verifier = UniqueContentVerifier(headless=True)
    
#     print verifier.is_content_unique("from all you maggots voting last week! Now it")
#     print verifier.is_content_unique("JIMMY PAGE Loses Planning Dispute With Neighbour")
#     print verifier.is_content_unique("sideman career with our comprehensive playlist, featuring collaborations")
#     print verifier.is_content_unique("This is a content summary only. Visit my website for dafasd  dafa")

    filename = 'reviews.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    for index in range(2,7222):
        ws.cell(row = index, column = 4).value = verifier.is_content_unique(ws.cell(row = index, column = 3).value.encode("utf8"))
        print index,ws.cell(row = index, column = 4).value
        wb.save(filename)
 
    verifier.close()
#     print RemovePunctuation("1,2,3,4,5")


if __name__ == '__main__':
    main()

